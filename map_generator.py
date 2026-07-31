import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
import webbrowser
import os
import folium

# поиск папки, в которой лежит скрипт
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# переносимые пути к файлам в папке "data"
EXCEL_PATH = os.path.join(BASE_DIR, "data", "rating_omsu_mock.xlsx")
GEOJSON_PATH = os.path.join(BASE_DIR, "data", "moscow_region.geojson")

# относительный путь к таблице
wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb["1_Ранг лучших динамика"]

data = []
for row in ws.iter_rows(min_row=10, max_col=16):
    omsu = row[3].value
    if omsu is None: continue
    data.append({
        "OMSU": str(omsu).strip(),
        "val_j": row[9].value,
        "val_p": row[15].value
    })
df = pd.DataFrame(data)

# --- 2) ЗАГРУЗКА ГРАНИЦ И ОБЪЕДИНЕНИЕ ---
gdf = gpd.read_file(GEOJSON_PATH)


def clean_name(name):
    if name is None: return ""
    name = str(name).lower()
    for word in ["г.о.", "городской округ", "г.", "область", "район"]:
        name = name.replace(word, "")
    return name.strip()


df["OMSU_clean"] = df["OMSU"].apply(clean_name)
gdf["district_clean"] = gdf["district"].apply(clean_name)

merged = gdf.merge(df, left_on="district_clean", right_on="OMSU_clean", how="left")

merged["val_j"] = merged["val_j"].astype(float).round(2)
merged["val_p"] = merged["val_p"].astype(float).round(2)

merged = merged.rename(columns={
    "district": "Округ",
    "val_j": "Отключения",
    "val_p": "Жалобы"
})

# --- 3) СОЗДАНИЕ КАРТЫ С ПУСТЫМ ФОНОМ И СКРИПТ УДАЛЕНИЯ ЛИШНЕГО ---
m = folium.Map(location=[55.75, 37.62], zoom_start=8, tiles=None, control_scale=False)

# CSS для меню + JS скрипт, который ВЫРЕЗАЕТ флаг и атрибуцию
custom_logic = """
<style>
    path:focus { outline: none !important; }
    .leaflet-interactive:focus { outline: none !important; }

    /* скрыть лишнее */
    .leaflet-control-attribution { display: none !important; visibility: hidden !important; }
    .leaflet-attribution-flag { display: none !important; visibility: hidden !important; }

    /* дизайн меню */
    .leaflet-control-layers {
        border: none !important;
        box-shadow: 0 6px 25px rgba(0,0,0,0.3) !important;
        border-radius: 15px !important;
        background-color: #ffffff !important;
        padding: 15px !important;
        font-family: 'Segoe UI', sans-serif !important;
        min-width: 220px;
    }
    .leaflet-control-layers-list::before {
        content: 'ВЫБОР СЛОЯ';
        font-weight: 900; display: block; font-size: 16px; color: #2c3e50;
        margin-bottom: 12px; letter-spacing: 2px; border-bottom: 3px solid #3498db; padding-bottom: 8px;
    }
    .leaflet-control-layers-base label {
        font-size: 17px !important; font-weight: 600 !important; color: #34495e !important;
        margin-bottom: 12px !important; cursor: pointer; display: flex; align-items: center;
    }
    .leaflet-control-layers-selector { transform: scale(1.6); margin-right: 15px !important; }
</style>

<script>
    // удаление лишнего
    window.onload = function() {
        var elements = document.getElementsByClassName('leaflet-control-attribution');
        while(elements.length > 0){
            elements[0].parentNode.removeChild(elements[0]);
        }
        var flags = document.getElementsByClassName('leaflet-attribution-flag');
        while(flags.length > 0){
            flags[0].parentNode.removeChild(flags[0]);
        }
    };
</script>
"""
m.get_root().header.add_child(folium.Element(custom_logic))

t_kwds = {
    "style": "font-size: 18px; font-weight: bold; background-color: white; border: 2px solid black; padding: 5px;"}

# --- 4) ОТРИСОВКА СЛОЕВ ---
merged.explore(
    m=m, column="Отключения", cmap="RdYlGn_r", tooltip=["Округ", "Отключения"],
    tooltip_kwds=t_kwds, name="Отключения 2025", popup=False,
    highlight=True, overlay=False,
    style_kwds=dict(fillOpacity=1.0, opacity=1.0, color="black", weight=1.2)
)

merged.explore(
    m=m, column="Жалобы", cmap="RdYlGn_r", tooltip=["Округ", "Жалобы"],
    tooltip_kwds=t_kwds, name="Жалобы 2025", popup=False,
    highlight=True, overlay=False,
    style_kwds=dict(fillOpacity=1.0, opacity=1.0, color="black", weight=1.2)
)

# --- 5) ТЮНИНГ ПОДСВЕТКИ ---
for child in m._children:
    if child.startswith('geo_json'):
        m._children[child].style_function = lambda x: {
            'fillColor': x['properties']['__folium_color'],
            'color': 'black', 'fillOpacity': 1.0, 'opacity': 1.0, 'weight': 1.2
        }
        m._children[child].highlight_function = lambda x: {
            'color': 'white', 'weight': 4.5, 'fillOpacity': 1.0
        }

folium.LayerControl(collapsed=False, position='topright').add_to(m)

# --- 6) СОХРАНЕНИЕ ---
path_file = os.path.join(BASE_DIR, "OMSU_Map_v2.html")
m.save(path_file)
webbrowser.open(f"file://{path_file}")
